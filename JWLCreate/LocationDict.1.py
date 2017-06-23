# 1. Create list of rows from spreadsheet
# 2. Create 'Location specific' list with no duplicates
# 3. Use the above to populate the 'Location' table in the userData.db
# 4. Populate the 'Note' table using information from thespreadsheet and
#       ensure correct 'LocationId' based on lookup of newly populated 
#       'Location' table

from openpyxl import load_workbook
import sqlite3
import uuid
from pytz import timezone
from datetime import datetime

### function to retrieve next usable 'locationId' from 'Locations' table in userData.db
def get_next_usable_location_id(sqlite3_connection):
    cursor = sqlite3_connection.cursor()
    for value_q1 in cursor.execute('SELECT COUNT(*) from Location'):
        if value_q1[0] != 0:
            location_ids = []
            for value_q2 in cursor.execute('SELECT LocationId FROM Location'):
                location_ids += value_q2
            next_usable_location_id = max(location_ids) + 1
            return next_usable_location_id
        else:
            return int(1)
    cursor.close()

#wb = load_workbook('D:\GitHub\Python_Scripts/Notes-3-Entries.xlsx')
wb = load_workbook('/home/dirk/GitHub/fluffy-couscous/JWLCreate/Test Data/Notes-3-Entries.xlsx')
sheetname = "Sheet1"
sheet = wb[sheetname]
 
bible_book_lookup = {'Genesis':'1', 'Exodus':'2', 'Leviticus':'3', 'Numbers':'4', 
'Deuteronomy':'5', 'Joshua':'6', 'Judges':'7', 'Ruth':'8', 
'1 Samuel':'9', '2 Samuel':'10', '1 Kings':'11', '2 Kings':'12', 
'1 Chronicles':'13', '2 Chronicles':'14', 'Ezra':'15', 'Nehemiah':'16', 
'Esther':'17', 'Job':'18', 'Psalms':'19', 'Proverbs':'20', 
'Ecclesiastes':'21', 'Song of Solomon':'22', 'Isaiah':'23', 'Jeremiah':'24', 
'Lamentations':'25', 'Ezekiel':'26', 'Daniel':'27', 'Hosea':'28', 
'Joel':'29', 'Amos':'30', 'Obadiah':'31', 'Jonah':'32', 
'Micah':'33', 'Nahum':'34', 'Habakkuk':'35', 'Zephaniah':'36', 
'Haggai':'37', 'Zechariah':'38', 'Malachi':'39', 
'Matthew':'40', 'Mark':'41', 'Luke':'42', 'John':'43', 
'Acts':'44', 'Romans':'45', '1 Corinthians':'46', '2 Corinthians':'47', 
'Galatians':'48', 'Ephesians':'49', 'Philippians':'50', 'Colossians':'51', 
'1 Thessalonians':'52', '2 Thessalonians':'53', '1 Timothy':'54', '2 Timothy':'55', 
'Titus':'56', 'Philemon':'57', 'Hebrews':'58', 'James':'59', 
'1 Peter':'60', '2 Peter':'61', '1 John':'62', '2 John':'63', 
'3 John':'64', 'Jude':'65', 'Revelation':'66'}

query_statement_start = "INSERT INTO Location(LocationId,BookNumber,ChapterNumber,DocumentId,Track,IssueTagNumber,KeySymbol,MepsLanguage,Type,Title) VALUES("
query_statement_end = ");"

### Create List of ALL rows from spreadsheet 
Row_List_With_Dups = []
for row in sheet.iter_rows():
    Row_List_With_Dups_Row = []
    for cell in row:
        Row_List_With_Dups_Row.append(cell.value)
    Row_List_With_Dups.append(Row_List_With_Dups_Row)
#print Row_List_With_Dups

### Create Location specific list (could have have duplicates)
Location_List_With_Dups = []
for row in Row_List_With_Dups:
    Location_List_With_Dups.append(row[:4])

### Remove duplicates in Location_List_With_Dups by creating new List
Location_List_Without_Dups = []
for row in Location_List_With_Dups:
    if row not in Location_List_Without_Dups:
        Location_List_Without_Dups.append(row[:4])
#print Location_List_Without_Dups

### Craft Location query strings
location_query_statement_start = "INSERT INTO Location(LocationId,BookNumber,ChapterNumber,DocumentId,Track,IssueTagNumber,KeySymbol,MepsLanguage,Type,Title) VALUES("
location_query_statement_end = ");"
sql_connection = sqlite3.Connection('/home/dirk/GitHub/fluffy-couscous/JWLCreate/Test Data/Truely-EMPTY/userData.db')
sql_connection_cursor=sql_connection.cursor()
location_id = get_next_usable_location_id(sql_connection)
for row in Location_List_Without_Dups:
    location_query_string = (
        location_query_statement_start +
        str(location_id) + "," +
        bible_book_lookup[row[0]] + "," + 
        str(row[1]) + "," +
        'NULL,NULL,0,"nwt",0,0,"' +
        (row[0] + " " + str(row[1])) + '"' +
        location_query_statement_end
    )
    location_id += 1
    print location_query_string

### Execute SQL query against userData.db
    sql_connection_cursor.execute(location_query_string)
    sql_connection.commit()

### Craft Note query strings
note_query_statement_start = "INSERT INTO Note(NoteId,Guid,UserMarkId,LocationId,Title,Content,LastModified,BlockType,BlockIdentifier) VALUES("
note_query_statement_end = ");"
note_id_start = 1
for row in Row_List_With_Dups:
    location_id_query_string = (str(row[0] + " " + str(row[1])),)
    sql_connection_cursor.execute('select LocationId from Location where Title=?', location_id_query_string)
    note_loction_id = sql_connection_cursor.fetchone()
    note_query_string = (
        note_query_statement_start +
        str(note_id_start) + ',"' +
        str(uuid.uuid4()) + '",' +
        "NULL" + "," +
        str(note_loction_id[0]) + ',"' +
        row[3] + '","' +
        row[4] + '","' +
        (timezone('GMT').localize(datetime.now()).replace(microsecond=0)).isoformat() + '",' +
        "2" + "," +
        str(row[2]) +
        note_query_statement_end
    )
    note_id_start += 1
    print note_query_string
### Execute SQL query to populate Note table
    sql_connection_cursor.execute(note_query_string)
    sql_connection.commit()

### Tidy up SQL cursor/connection
sql_connection_cursor.close()
sql_connection.close()
