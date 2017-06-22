from openpyxl import load_workbook

#wb = load_workbook('G:/GitHub/fluffy-couscous/JWLCreate/Test Data/Notes-3-Entries.xlsx')
wb = load_workbook('/home/dirk/GitHub/fluffy-couscous/JWLCreate/Test Data/Notes-3-Entries.xlsx')
sheetname = "Sheet1"
sheet = wb[sheetname]

bible_book_lookup = {'Genesis':'1', 'Exodus':'2', 'Leviticus':'3', 'Numbers':'4', 'Deuteronomy':'5', 'Joshua':'6', 'Judges':'7', 'Ruth':'8', '1 Samuel':'9', '2 Samuel':'10', '1 Kings':'11', '2 Kings':'12', '1 Chronicles':'13', '2 Chronicles':'14', 'Ezra':'15', 'Nehemiah':'16', 'Esther':'17', 'Job':'18', 'Psalms':'19', 'Proverbs':'20', 'Ecclesiastes':'21', 'Song of Solomon':'22', 'Isaiah':'23', 'Jeremiah':'24', 'Lamentations':'25', 'Ezekiel':'26', 'Daniel':'27', 'Hosea':'28', 'Joel':'29', 'Amos':'30', 'Obadiah':'31', 'Jonah':'32', 'Micah':'33', 'Nahum':'34', 'Habakkuk':'35', 'Zephaniah':'36', 'Haggai':'37', 'Zechariah':'38', 'Malachi':'39', 'Matthew':'40', 'Mark':'41', 'Luke':'42', 'John':'43', 'Acts':'44', 'Romans':'45', '1 Corinthians':'46', '2 Corinthians':'47', 'Galatians':'48', 'Ephesians':'49', 'Philippians':'50', 'Colossians':'51', '1 Thessalonians':'52', '2 Thessalonians':'53', '1 Timothy':'54', '2 Timothy':'55', 'Titus':'56', 'Philemon':'57', 'Hebrews':'58', 'James':'59', '1 Peter':'60', '2 Peter':'61', '1 John':'62', '2 John':'63', '3 John':'64', 'Jude':'65', 'Revelation':'66'}
location_id = 1
query_start = "INSERT INTO Location(LocationId,BookNumber,ChapterNumber,DocumentId,Track,IssueTagNumber,KeySymbol,MepsLanguage,Type,Title) VALUES("
query_end = "');"
total_query_values = []
check_for_dups = []

for row in sheet.iter_rows():
    listRow = []
    for cell in row:
        listRow.append(cell.value)
        book_number = bible_book_lookup[listRow[0]]
    query_values = [
#        str(location_id), 
        str(book_number), 
        str(listRow[1]), 
        "NULL", 
        "NULL", 
        "0", 
        "nwt", 
        "0", 
        "0",
        (str(listRow[0]) + " " + str(listRow[1]))
    ]
    location_id = location_id + 1
    if query_values not in total_query_values:
        total_query_values.append(query_values)

print (total_query_values)

# closest but Location_id missing
