from openpyxl import load_workbook
import uuid
from time import gmtime, strftime
from datetime import datetime
from pytz import timezone
from datetime import datetime


wb = load_workbook('G:/GitHub/fluffy-couscous/JWLCreate/Test Data/Notes-3-Entries.xlsx')
sheetname = "Sheet1"
sheet = wb[sheetname]
notesQueryStart = "INSERT INTO Note(NoteId,Guid,UserMarkId,LocationId,Title,Content,LastModified,BlockType,BlockIdentifier) VALUES("
notesQueryEnd = ");"
noteIdStart = 1

for row in sheet.iter_rows():
    listRow = []
    for cell in row:
        listRow.append(cell.value)
        
    print (
        notesQueryStart + 
        str(noteIdStart) + "," +
        str(uuid.uuid4()) + "," +
        "NULL" + "," +
        "LOCATIONID" + "," +
        listRow[3] + "," +
        listRow[4] + "," +
        (timezone('GMT').localize(datetime.now()).replace(microsecond=0)).isoformat() + "," +
        "2" + "," +
        str(listRow[2]) +
        notesQueryEnd
        )
    noteIdStart = noteIdStart + 1
    